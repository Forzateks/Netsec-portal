"""
Phase 3 task migration preprocessor (v96).
Reads Daily_Tracker-Internal.xlsx, normalizes, emits migration_plan.json + validation_report.txt.

Patches applied vs the original spec script (per Nasif's review):
  - Issue 1: Not Completed shifted-column branch fixed — title is at row[1]
             in BOTH layouts; shifted rows match the General Tasks schema
             (10 cols) and carry a priority at row[2].
  - Issue 2: task_templates has no `status` column in schema → dropped
             from add_template() / templates_out.
  - Issue 3: task_templates HAS a priority column with DB default 'medium';
             recurring sheets have no priority col, so we omit it and the
             default applies on INSERT.
  - Issue 4: parse_date() now records dirty values into DIRTY_DATES with
             context, surfaced in the validation report.
  - Issue 6: within-sheet duplicate titles detected + reported (not auto-
             skipped — legitimate repeats are possible).
"""
import json, re
from datetime import datetime, timedelta
from collections import Counter
from openpyxl import load_workbook
from pathlib import Path

SRC = Path(__file__).parent / "Daily_Tracker-Internal.xlsx"
OUT_JSON = Path(__file__).parent / "migration_plan.json"
OUT_REPORT = Path(__file__).parent / "validation_report.txt"

# ---------- Normalization maps ----------
NAME_MAP = {
    "venkat": "Venkatesan",
    "venkatesan": "Venkatesan",
    "ahmed": "Ahmed Ali",
    "ahmed ali": "Ahmed Ali",
    "nasif": "Mohammed Nasif",
    "mohammed nasif": "Mohammed Nasif",
    "afsal": "Mohammed Afsal",
    "mohammed afsal": "Mohammed Afsal",
    "salman": "Salman Aziz",
    "salman aziz": "Salman Aziz",
    "prasanth": "Prasanth",
}

STATUS_MAP = {
    "completed": "completed",
    "ongoing": "ongoing",
    "on going": "ongoing",
    "ongping": "ongoing",  # typo in source
    "yet to start": "yet_to_start",
}

PRIORITY_MAP = {"high": "high", "medium": "medium", "low": "low"}

# ---------- Helpers ----------
DIRTY_DATES = []  # collected for the validation report

def clean(v):
    if v is None: return None
    s = str(v).strip()
    return s if s else None

def normalize_status(raw, default):
    s = clean(raw)
    if not s: return default
    return STATUS_MAP.get(s.lower(), default)

def normalize_priority(raw):
    s = clean(raw)
    if not s: return "medium"
    return PRIORITY_MAP.get(s.lower(), "medium")

def parse_owners(raw):
    """Returns (resolved, unresolved). Both lists may be empty."""
    s = clean(raw)
    if not s or s.upper() == "NA": return [], []
    tokens = [t.strip() for t in re.split(r"[/,]", s) if t.strip()]
    resolved, unresolved = [], []
    for t in tokens:
        key = t.lower().strip()
        if key in NAME_MAP:
            resolved.append(NAME_MAP[key])
        else:
            unresolved.append(t)
    return resolved, unresolved

def parse_date(v, context=None):
    """ISO date or None. Junk values are appended to DIRTY_DATES."""
    if v is None: return None
    if isinstance(v, datetime): return v.date().isoformat()
    if isinstance(v, (int, float)):
        try:
            return (datetime(1899, 12, 30) + timedelta(days=int(v))).date().isoformat()
        except Exception:
            DIRTY_DATES.append((context, v))
            return None
    s = str(v).strip()
    if not s or s.upper() == "NA": return None
    try:
        return datetime.fromisoformat(s).date().isoformat()
    except Exception:
        DIRTY_DATES.append((context, s))
        return None

# ---------- Extractors ----------
wb = load_workbook(SRC, data_only=True)
issues = []
tasks_out = []
templates_out = []

def add_task(source, title, desc, priority, status, start, eta, end, owners_raw, remarks, frequency):
    if not title:
        issues.append(f"[{source}] row skipped - no title")
        return
    owners, unresolved = parse_owners(owners_raw)
    if unresolved:
        issues.append(f"[{source}] unresolved owner(s) for '{title[:60]}': {unresolved}")
    ctx = f"{source}/{title[:30]}"
    tasks_out.append({
        "source_sheet": source,
        "title": title.strip(),
        "description": clean(desc),
        "priority": normalize_priority(priority),
        "status": status,
        "frequency": frequency,
        "start_date": parse_date(start, ctx + "/start"),
        "eta_date":   parse_date(eta,   ctx + "/eta"),
        "end_date":   parse_date(end,   ctx + "/end"),
        "remarks": f"[MIG-v96] {clean(remarks) or ''}".strip(),
        "assignees": owners,
    })

def add_template(source, title, desc, freq, owners_raw, remarks):
    # No status / no priority — schema has priority with DB default 'medium'
    # and no status column at all. Both are intentionally omitted.
    if not title:
        issues.append(f"[{source}] template row skipped - no title")
        return
    owners, unresolved = parse_owners(owners_raw)
    if unresolved:
        issues.append(f"[{source}] unresolved template owner(s) for '{title[:60]}': {unresolved}")
    templates_out.append({
        "source_sheet": source,
        "title": title.strip(),
        "description": clean(desc),
        "frequency": freq,
        "remarks": f"[MIG-v96] {clean(remarks) or ''}".strip(),
        "assignees": owners,
    })

# General Tasks -> tasks, frequency=general
# Columns: Task, Priority, Description, Start Date, ETA, End date, Owner, Status, Remarks
ws = wb["General Tasks"]
for row in ws.iter_rows(min_row=2, values_only=True):
    if not any(row): continue
    task, prio, desc, start, eta, end, owner, status, remarks = row[:9]
    add_task("General Tasks", clean(task), desc, prio,
             normalize_status(status, "yet_to_start"),
             start, eta, end, owner, remarks, "general")

# General Tasks - Completed -> tasks, status forced to completed
# Columns: Description (=title), Priority, Descripton[typo, real desc], Start Date, ETA, End Dtae, Owner, Status, Remarks
# Sheet has its header on row 3 (dims=A3:J...); real data starts row 4.
ws = wb["General Tasks - Completed"]
for row in ws.iter_rows(min_row=4, values_only=True):
    if not any(row): continue
    title, prio, desc, start, eta, end, owner, status, remarks = row[:9]
    add_task("Completed", clean(title), desc, prio,
             "completed",  # forced per decision #1
             start, eta, end, owner, remarks, "general")

# General Tasks - Not Completed -> tasks, status=ongoing default
# Two layouts coexist on this sheet:
#   Standard (8 cols): SN, title, desc, start, end, owner, status, remarks
#   Shifted  (10 cols): SN, title, priority, desc, start, eta, end, owner, status, remarks
# Detection: row[2] parses as a priority value. Title is at row[1] in both.
# Sheet has its header on row 3 (dims=A3:J...); real data starts row 4.
ws = wb["General Tasks - Not Completed"]
for row in ws.iter_rows(min_row=4, values_only=True):
    if not any(row): continue
    col2 = clean(row[2]) if len(row) > 2 else None
    if col2 and col2.lower() in ("high", "medium", "low"):
        title   = clean(row[1])
        prio    = row[2]
        desc    = row[3] if len(row) > 3 else None
        start   = row[4] if len(row) > 4 else None
        eta     = row[5] if len(row) > 5 else None
        end     = row[6] if len(row) > 6 else None
        owner   = row[7] if len(row) > 7 else None
        status  = row[8] if len(row) > 8 else None
        remarks = row[9] if len(row) > 9 else None
    else:
        _, title, desc, start, end, owner, status, remarks = row[:8]
        title = clean(title)
        prio = None
        eta = None
    add_task("Not Completed", title, desc, prio,
             normalize_status(status, "ongoing"),
             start, eta, end, owner, remarks, "general")

# Daily / Weekly / Monthly / Quarterly -> task_templates
# Columns: SN, Task, Description, Frequency, Start Date, End Dtae, Status, Remarks, Owner
SHEET_TO_FREQ = {"Daily": "daily", "Weekly": "weekly", "Monthly": "monthly", "Quarterly": "quarterly"}
for sheet_name, freq in SHEET_TO_FREQ.items():
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=4, values_only=True):  # data starts row 4 per inspection
        if not any(row): continue
        sn, task, desc, _, _, _, status, remarks, owner = row[:9]
        if not clean(task): continue
        # status intentionally not propagated — schema has no column
        add_template(sheet_name, clean(task), desc, freq, owner, remarks)

# ---------- Within-sheet duplicate detection (issue 6) ----------
dup_counter = Counter((t["source_sheet"], t["title"]) for t in tasks_out)
within_sheet_dups = {k: v for k, v in dup_counter.items() if v > 1}

# ---------- Write outputs ----------
plan = {"tasks": tasks_out, "templates": templates_out}
OUT_JSON.write_text(json.dumps(plan, indent=2, default=str), encoding="utf-8")

with OUT_REPORT.open("w", encoding="utf-8") as f:
    f.write(f"Tasks to insert:     {len(tasks_out)}\n")
    f.write(f"Templates to insert: {len(templates_out)}\n")
    f.write(f"\nBy source sheet:\n")
    for k, v in Counter(t["source_sheet"] for t in tasks_out).items():
        f.write(f"  {k}: {v}\n")
    for k, v in Counter(t["source_sheet"] for t in templates_out).items():
        f.write(f"  {k}: {v}\n")
    f.write(f"\nTasks with zero assignees: {sum(1 for t in tasks_out if not t['assignees'])}\n")
    f.write(f"Tasks with multi-assignee: {sum(1 for t in tasks_out if len(t['assignees']) > 1)}\n")

    f.write(f"\nWithin-sheet duplicate titles: {len(within_sheet_dups)}\n")
    for (sheet, title), count in within_sheet_dups.items():
        f.write(f"  - [{sheet}] '{title}' x{count}\n")

    empty_templates = [t["title"] for t in templates_out if not t["assignees"]]
    f.write(f"\n⚠️  Templates with ZERO assignees ({len(empty_templates)}):\n")
    for t in empty_templates:
        f.write(f"  - {t}\n")
    f.write("  → Edit in UI before next generateMissingInstances run, "
            "or empty-owner instances will recur every period.\n")

    f.write(f"\nDirty dates dropped to NULL: {len(DIRTY_DATES)}\n")
    for ctx, val in DIRTY_DATES:
        f.write(f"  - {ctx}: {val!r}\n")

    f.write(f"\nIssues ({len(issues)}):\n")
    for i in issues:
        f.write(f"  - {i}\n")

print(f"Wrote {OUT_JSON} and {OUT_REPORT}")
print(f"Tasks: {len(tasks_out)}  Templates: {len(templates_out)}  Issues: {len(issues)}  DirtyDates: {len(DIRTY_DATES)}  Dups: {len(within_sheet_dups)}")
